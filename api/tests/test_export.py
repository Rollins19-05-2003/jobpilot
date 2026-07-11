import io
import shutil
import subprocess

import pytest
from openpyxl import load_workbook

from app.export.xlsx import COMPANY_HEADERS, CONTACT_HEADERS, EXPECTED_SHEETS, verify_workbook


def _export(client) -> bytes:
    client.post("/companies/seed")
    cid = client.post("/contacts", json={
        "company": "Razorpay", "first_name": "Priya", "persona": "recruiter", "hook": "payments team",
    }).json()["id"]
    client.patch(f"/contacts/{cid}", json={"status": "sent"})
    r = client.get("/export/tracker.xlsx")
    assert r.status_code == 200
    return r.content


def test_workbook_has_four_sheets_with_expected_headers(client):
    data = _export(client)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == EXPECTED_SHEETS

    companies = wb["🏢 Company Tracker"]
    assert [c.value for c in companies[2]] == COMPANY_HEADERS
    assert companies.max_row == 2 + 50  # all 50 seeded companies

    outreach = wb["✉️ Outreach Tracker"]
    assert [c.value for c in outreach[2]] == CONTACT_HEADERS
    assert outreach["B4"].value == "Priya"                      # row 3 is the sample row
    assert outreach.cell(row=4, column=12).value.startswith("=IF(K4")  # Days Since Sent
    assert outreach["H1"].value == "=IFERROR(F1/D1,0)"          # reply-rate guard
    assert outreach["H1"].number_format == "0.0%"


def test_workbook_passes_broken_formula_scan(client):
    verify_workbook(_export(client))  # raises on #REF!/#NAME? or wrong sheets


def test_export_contains_no_email_sending_machinery(client):
    # Ethics guard: the workbook teaches manual sending; the API must not send.
    from app.main import app
    paths = {r.path for r in app.routes}
    assert not any("send" in p for p in paths if p.startswith("/outreach"))


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice not installed")
def test_libreoffice_recalc(client, tmp_path):
    path = tmp_path / "tracker.xlsx"
    path.write_bytes(_export(client))
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(tmp_path / "out"), str(path)],
        check=True, timeout=120,
    )
    recalced = load_workbook(tmp_path / "out" / "tracker.xlsx", data_only=True)
    for ws in recalced.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.startswith("#")), \
                    f"{ws.title}!{cell.coordinate} = {cell.value}"
