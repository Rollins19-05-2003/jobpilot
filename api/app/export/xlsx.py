"""GET /export/tracker.xlsx — the full outreach workbook, generated fresh from
the live DB on every call (openpyxl, in-memory, nothing written to disk).

Four tabs: 📖 README (how to run the system), 🏢 Company Tracker (tiered
targets), ✉️ Outreach Tracker (contacts + live formulas), 📧 Email Templates.
After building, the workbook is re-opened and scanned for broken formulas
(#REF!/#NAME?) before it ever leaves the API.
"""
import io
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..db import Company, Contact, SessionLocal

log = logging.getLogger("jobpilot.export")

router = APIRouter(tags=["export"])

# ------------------------------------------------------------------ palette

SLATE = "1F2937"          # header rows
GREY_FILL = "F3F4F6"      # email bodies
SAMPLE_YELLOW = "FFF9C4"  # the sample contact row
TIER_FILLS = {1: "FFF9E6", 2: "E8F1FB", 3: "E7F6EC", 4: "FDECF1", 5: "ECEBFA"}

ARIAL = "Arial"
COURIER = "Courier New"

HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=SLATE)
TITLE_FONT = Font(name=ARIAL, bold=True, size=16, color=SLATE)
H2_FONT = Font(name=ARIAL, bold=True, size=12, color=SLATE)
BODY_FONT = Font(name=ARIAL, size=10)
BOLD_FONT = Font(name=ARIAL, size=10, bold=True)
MONO_FONT = Font(name=COURIER, size=10)
LINK_FONT = Font(name=ARIAL, size=10, color="0563C1", underline="single")
WARN_FONT = Font(name=ARIAL, size=10, bold=True, color="B91C1C")

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _no_gridlines(ws) -> None:
    ws.sheet_view.showGridLines = False


def _set_widths(ws, widths: dict[int, int]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def _header_row(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _link_cell(cell, url: str, label: str) -> None:
    cell.value = label
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT
    else:
        cell.font = BODY_FONT


# ------------------------------------------------------------------- README

README_SECTIONS: list[tuple[str, list[str]]] = [
    ("How to use this workbook", [
        "🏢 Company Tracker — your 50 tiered targets. Work top tier + high priority first. "
        "The LinkedIn Search column is a pre-built people-search for that company's recruiters.",
        "✉️ Outreach Tracker — one row per human you contact. The API fills this from the DB; "
        "'Days Since Sent' updates itself every time you open the file.",
        "📧 Email Templates — the three message shapes + five ready hooks. Copy, personalise, send.",
        "This file is a SNAPSHOT — regenerate anytime with: curl -o tracker.xlsx http://localhost:8000/export/tracker.xlsx",
    ]),
    ("The daily 30-minute workflow", [
        "1. Read the Telegram digest (2 min) — new roles + follow-ups due today.",
        "2. Apply to any strong-fit roles it flagged (10 min).",
        "3. Pick 1–2 companies from Company Tracker (priority=high, status=not_started) (1 min).",
        "4. Open their LinkedIn Search link, pick 2–3 recruiters or eng managers (5 min).",
        "5. Find + verify their emails with the finder stack below (5 min).",
        "6. Add each as a contact with ONE specific hook, then generate a draft "
        "(POST /contacts → POST /outreach/draft/{id}) (3 min).",
        "7. Review the draft, personalise if needed, send from Gmail 9–11 AM recipient time (4 min).",
        "8. Mark it sent (PATCH /contacts/{id} status=sent). Follow-ups surface automatically after 5 days.",
    ]),
    ("Free email-finder stack (monthly free-tier limits)", [
        "Hunter.io — 25 searches/mo. Best for pattern discovery (first.last@company.com).",
        "Snov.io — 50 credits/mo. Good LinkedIn extension.",
        "RocketReach — 5 lookups/mo. Save for high-value contacts.",
        "Apollo.io — free tier with limited email credits/mo; decent people search.",
        "FindThatLead — small trial credits.",
        "Mailtester.com / verifalia — free single-address verification before you send. ALWAYS verify.",
    ]),
    ("Mail-merge tools (free daily limits)", [
        "YAMM (Yet Another Mail Merge) — ~50 emails/day free, lives inside Google Sheets.",
        "GMass — free tier, Gmail-native.",
        "Mailmeteor — 50/day free.",
        "All three read a Google Sheet — paste rows straight from the Outreach Tracker tab.",
    ]),
    ("⚠️ WARNINGS — deliverability and sanity", [
        "NEVER BCC a list at scale. One personalised email per human, always.",
        "Warm up a fresh inbox: max 10 sends/day for the first 3–4 days.",
        "No links, no attachments in the FIRST email — they trip spam filters.",
        "Send 9–11 AM in the recipient's timezone (that's IST for Bangalore targets).",
        "If reply rate is 0% after 30 sends: STOP. The hook is broken — rewrite it. Volume never fixes message.",
        "JobPilot drafts; YOU send. Review every single email before it leaves. No automated sending, ever.",
    ]),
]


def _build_readme(ws) -> None:
    ws.title = "📖 README"
    _no_gridlines(ws)
    _set_widths(ws, {1: 3, 2: 120})
    ws.cell(row=2, column=2, value="🛩️ JobPilot — Outreach Tracker").font = TITLE_FONT
    sub = ws.cell(
        row=3, column=2,
        value=f"Cold outreach, tracked like a pipeline. Generated {datetime.now(timezone.utc):%d %b %Y %H:%M} UTC "
              "from the live JobPilot DB.",
    )
    sub.font = Font(name=ARIAL, size=10, italic=True, color="6B7280")
    row = 5
    for heading, lines in README_SECTIONS:
        cell = ws.cell(row=row, column=2, value=heading)
        cell.font = WARN_FONT if heading.startswith("⚠️") else H2_FONT
        row += 1
        for line in lines:
            c = ws.cell(row=row, column=2, value=line)
            c.font = BODY_FONT
            c.alignment = WRAP
            row += 1
        row += 1


# --------------------------------------------------------- company tracker

COMPANY_HEADERS = ["#", "Company", "Tier", "HQ", "Why Target", "LinkedIn Search",
                   "Careers", "Priority", "Status", "People Found", "Emails Sent", "Notes"]


def _build_companies(ws, companies: list[Company], contact_counts: dict, sent_counts: dict) -> None:
    ws.title = "🏢 Company Tracker"
    _no_gridlines(ws)
    _set_widths(ws, {1: 4, 2: 22, 3: 6, 4: 24, 5: 52, 6: 16, 7: 16, 8: 9, 9: 13, 10: 13, 11: 12, 12: 30})

    first_data, last_data = 3, 2 + len(companies)
    # Live counter strip (row 1) — formulas over the table range, so the sheet
    # stays honest even when edited by hand later.
    counters = [
        ("Companies", f"=COUNTA(B{first_data}:B{last_data})"),
        ("People found", f"=SUM(J{first_data}:J{last_data})"),
        ("Emails sent", f"=SUM(K{first_data}:K{last_data})"),
        ("Outreaching", f'=COUNTIF(I{first_data}:I{last_data},"outreaching")'),
        ("Done", f'=COUNTIF(I{first_data}:I{last_data},"done")'),
    ]
    col = 2
    for label, formula in counters:
        ws.cell(row=1, column=col, value=label).font = BOLD_FONT
        fc = ws.cell(row=1, column=col + 1, value=formula)
        fc.font = Font(name=ARIAL, size=10, bold=True, color="0563C1")
        col += 2

    _header_row(ws, 2, COMPANY_HEADERS)
    for i, c in enumerate(companies):
        r = first_data + i
        fill = PatternFill("solid", fgColor=TIER_FILLS.get(c.tier, "FFFFFF"))
        values = [i + 1, c.name, f"T{c.tier}", c.hq_location, c.why_target,
                  None, None, c.priority, c.status,
                  contact_counts.get(c.id, 0), sent_counts.get(c.id, 0), c.notes or ""]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx)
            if col_idx == 6:
                _link_cell(cell, c.linkedin_people_url, "Recruiters →")
            elif col_idx == 7:
                _link_cell(cell, c.careers_url, "Careers →")
            else:
                cell.value = v
                cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BORDER
            if col_idx in (5, 12):
                cell.alignment = WRAP
            if col_idx in (1, 3, 8, 9, 10, 11):
                cell.alignment = CENTER
    ws.freeze_panes = "B3"  # header + counters stay put; first column pinned


# --------------------------------------------------------- outreach tracker

CONTACT_HEADERS = ["#", "First Name", "Last Name", "Company", "Role Title", "Persona",
                   "LinkedIn", "Email", "Source", "Verified", "Date Sent", "Days Since Sent",
                   "Status", "Follow-up Sent", "Follow-up Date", "Hook", "Outcome", "Notes"]
PAD_TO = 100  # empty, pre-formatted rows so the sheet works as a manual tracker too


def _build_contacts(ws, contacts: list[Contact]) -> None:
    ws.title = "✉️ Outreach Tracker"
    _no_gridlines(ws)
    _set_widths(ws, {1: 4, 2: 12, 3: 12, 4: 18, 5: 22, 6: 12, 7: 14, 8: 28, 9: 10,
                     10: 9, 11: 12, 12: 10, 13: 11, 14: 10, 15: 12, 16: 44, 17: 20, 18: 24})

    # Data layout: row 2 = header, row 3 = yellow sample row, rows 4.. = real
    # contacts + padding. Stats formulas skip the sample row.
    first_data, last_data = 4, 3 + max(len(contacts), PAD_TO)
    stats = [
        ("Contacts", f"=COUNTA(B{first_data}:B{last_data})", None),
        ("Sent", f"=COUNT(K{first_data}:K{last_data})", None),
        ("Replies", f'=COUNTIF(M{first_data}:M{last_data},"replied")+COUNTIF(M{first_data}:M{last_data},"meeting")', None),
        ("Reply rate", "=IFERROR(F1/D1,0)", "0.0%"),  # F1=Replies, D1=Sent
    ]
    col = 1
    for label, formula, fmt in stats:
        ws.cell(row=1, column=col, value=label).font = BOLD_FONT
        fc = ws.cell(row=1, column=col + 1, value=formula)
        fc.font = Font(name=ARIAL, size=10, bold=True, color="0563C1")
        if fmt:
            fc.number_format = fmt
        col += 2

    _header_row(ws, 2, CONTACT_HEADERS)

    sample = ["·", "Priya", "Sharma", "Razorpay", "Senior TA Partner", "recruiter",
              "linkedin.com/in/…", "priya@razorpay.com", "hunter", "yes", None, None,
              "sent", "no", None, "Hiring for the payments platform team (JD mentions GCP + Python)",
              "", "SAMPLE ROW — copy this shape"]
    for col_idx, v in enumerate(sample, start=1):
        cell = ws.cell(row=3, column=col_idx, value=v)
        cell.font = BODY_FONT
        cell.fill = PatternFill("solid", fgColor=SAMPLE_YELLOW)
        cell.border = BORDER

    for i in range(max(len(contacts), PAD_TO)):
        r = first_data + i
        c = contacts[i] if i < len(contacts) else None
        if c is not None:
            values = [i + 1, c.first_name, c.last_name, c.company.name if c.company else "",
                      c.role_title, c.persona, c.linkedin_url, c.email, c.email_source,
                      "yes" if c.email_verified else "no",
                      c.date_sent.date() if c.date_sent else None, None,
                      c.status, "yes" if c.follow_up_sent else "no",
                      c.follow_up_date.date() if c.follow_up_date else None,
                      c.hook, c.outcome or "", c.notes or ""]
        else:
            values = [None] * len(CONTACT_HEADERS)
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER
            cell.font = BODY_FONT
            if col_idx == 12:
                # K = Date Sent → self-updating age; blank until a date exists
                cell.value = f'=IF(K{r}="","",TODAY()-K{r})'
                cell.alignment = CENTER
            elif v is not None:
                cell.value = v
                if col_idx in (11, 15):
                    cell.number_format = "DD-MMM-YY"
                if col_idx == 16:
                    cell.alignment = WRAP
                if col_idx in (1, 6, 9, 10, 13, 14):
                    cell.alignment = CENTER
    ws.freeze_panes = "B3"


# ---------------------------------------------------------- email templates

TEMPLATE_A_EXAMPLE = (
    "Subject: Backend engineer (Python/GCP) — {{Company}}\n"
    "\n"
    "Hi {{First Name}},\n"
    "\n"
    "I'm Sourav — backend engineer at Quantacus.ai in Bangalore (~1.5 yrs, Python/FastAPI/GCP).\n"
    "{{Hook}} — that's exactly the kind of problem I've been working on: my LLM pipelines\n"
    "process 2–5 crore records per brand across 20+ brands.\n"
    "\n"
    "Would you be open to pointing me at open backend/SDE roles at {{Company}} that might fit?\n"
    "\n"
    "Thanks,\nSourav"
)

TEMPLATE_B_EXAMPLE = (
    "Subject: How does {{Company}} handle feed-scale writes?\n"
    "\n"
    "Hi {{First Name}},\n"
    "\n"
    "{{Hook}} — I've been chewing on the same class of problem at Quantacus, where I rebuilt a\n"
    "catalog pipeline from 12,000s to 11s, and I'm curious how your team approaches it at\n"
    "{{Company}}'s scale: do you batch through a queue or stream per-record?\n"
    "\n"
    "If you have 15 minutes sometime, I'd genuinely value your advice on how engineers grow\n"
    "into these problems. No agenda beyond that.\n"
    "\n"
    "Thanks,\nSourav"
)

TEMPLATE_FOLLOWUP_EXAMPLE = (
    "Subject: Re: <original subject>\n"
    "\n"
    "Hi {{First Name}}, wrote to you last week about backend roles at {{Company}}.\n"
    "Since then I shipped JobPilot — a zero-cost FastAPI + Gemini + n8n pipeline that automates\n"
    "my job search end-to-end. If now's not the right time, no stress at all.\n"
    "\n"
    "Sourav"
)

WHY_B_WORKS = (
    "Why Template B out-performs Template A with engineering managers: it asks for nothing they "
    "must refuse (no role request → no reflex to forward you to HR), it flatters expertise — "
    "people enjoy being asked how they built something — and 15 minutes of advice is a far easier "
    "'yes' than 'refer this stranger'. The referral usually follows anyway once you've talked."
)

EXAMPLE_HOOKS = [
    "Saw you're hiring for the catalog/feeds team — I process 2–5 crore records per brand across 20+ brands at Quantacus, so your scale problems are literally my day job.",
    "Your JD mentions cost optimisation at scale — I cut per-SKU processing cost ~83% (Rs. 3 → Rs. 0.5) on production LLM pipelines.",
    "Noticed your team owns the data-pipeline layer — I recently redesigned one from 12,000+ seconds to 11 seconds and I'm still not over it.",
    "Read that you're scaling the experimentation platform — I shaved ~70% off API latency in a large-scale experimentation module last quarter.",
    "Your post about LLM infra hit home — I orchestrate Gemini/OpenAI pipelines in production on GCP (BigQuery, Cloud Run, Cloud Tasks) at Quantacus.",
]


def _build_templates(ws) -> None:
    ws.title = "📧 Email Templates"
    _no_gridlines(ws)
    _set_widths(ws, {1: 3, 2: 110})

    ws.cell(row=2, column=2, value="Email Templates — copy, personalise, send manually").font = TITLE_FONT
    ws.cell(row=3, column=2,
            value="Placeholders {{First Name}} / {{Company}} / {{Hook}} are for mail-merge. "
                  "The API's /outreach/draft endpoint generates fully-resolved versions of these.").font = BODY_FONT

    blocks = [
        ("Template A — recruiters / talent acquisition (<120 words, soft ask)", TEMPLATE_A_EXAMPLE),
        ("Template B — engineering managers (ask for advice, not a job)", TEMPLATE_B_EXAMPLE),
        ("Follow-up — 5+ days after no reply (2–3 lines, one NEW piece of value)", TEMPLATE_FOLLOWUP_EXAMPLE),
    ]
    row = 5
    for heading, body in blocks:
        ws.cell(row=row, column=2, value=heading).font = H2_FONT
        row += 1
        for line in body.split("\n"):
            c = ws.cell(row=row, column=2, value=line if line else " ")
            c.font = MONO_FONT
            c.fill = PatternFill("solid", fgColor=GREY_FILL)
            row += 1
        row += 1

    ws.cell(row=row, column=2, value="Why B beats A").font = H2_FONT
    c = ws.cell(row=row + 1, column=2, value=WHY_B_WORKS)
    c.font = BODY_FONT
    c.alignment = WRAP
    row += 3

    ws.cell(row=row, column=2, value="Five ready hooks (built from real metrics — adapt, don't paste)").font = H2_FONT
    row += 1
    for i, hook in enumerate(EXAMPLE_HOOKS, start=1):
        c = ws.cell(row=row, column=2, value=f"{i}. {hook}")
        c.font = BODY_FONT
        c.alignment = WRAP
        row += 1


# ------------------------------------------------------------ build + verify

def build_workbook() -> bytes:
    with SessionLocal() as s:
        companies = s.query(Company).order_by(Company.tier, Company.name).all()
        contacts = s.query(Contact).order_by(Contact.created_at).all()
        contact_counts: dict[int, int] = {}
        sent_counts: dict[int, int] = {}
        for c in contacts:
            contact_counts[c.company_id] = contact_counts.get(c.company_id, 0) + 1
            if c.date_sent:
                sent_counts[c.company_id] = sent_counts.get(c.company_id, 0) + 1

        wb = Workbook()
        _build_readme(wb.active)
        _build_companies(wb.create_sheet(), companies, contact_counts, sent_counts)
        _build_contacts(wb.create_sheet(), contacts)
        _build_templates(wb.create_sheet())

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


EXPECTED_SHEETS = ["📖 README", "🏢 Company Tracker", "✉️ Outreach Tracker", "📧 Email Templates"]


def verify_workbook(data: bytes) -> None:
    """Re-open what we just built and fail loudly on broken formulas/structure.
    (A LibreOffice headless recalc is a stronger check — run it in CI/tests if
    soffice is installed; openpyxl alone never evaluates formulas.)"""
    wb = load_workbook(io.BytesIO(data))
    if wb.sheetnames != EXPECTED_SHEETS:
        raise ValueError(f"unexpected sheets: {wb.sheetnames}")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and ("#REF!" in cell.value or "#NAME?" in cell.value):
                    raise ValueError(f"broken formula at {ws.title}!{cell.coordinate}: {cell.value}")


@router.get("/export/tracker.xlsx")
def export_tracker() -> StreamingResponse:
    try:
        data = build_workbook()
        verify_workbook(data)
    except Exception as e:  # noqa: BLE001 — surface build problems as a clean 500, not a stack trace
        log.error("tracker export failed: %s", e)
        raise HTTPException(500, f"tracker export failed: {e}")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="jobpilot_tracker.xlsx"'},
    )
